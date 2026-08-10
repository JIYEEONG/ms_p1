# app/services/excel_export.py
"""
매출리포트_샘플.xlsx 구조 기반 상세 엑셀 생성.
excel_report_builder.build_excel_report_data()의 결과를 받아 워크북을 만든다.

시트 구성:
  1. 판매성과리포트 - 대분류/중분류 Top5, 베스트10, 해당기간 vs 비교기간 이중꺾은선 네이티브 차트
  2. 해당기간 판매 상세 - 판매된 SKU 전체 (정상/취소/반품/교환 모두 포함)
  3. 비교기간 판매 상세 - 동일 구조

⚠ 목표매출/달성률: 현재 시스템 전체에서 목표매출 계산 로직이 미정(하드코딩 이슈 별도 확인 중)이라
   이 엑셀에서도 "-"로 비워둔다. 임의로 새 계산식을 만들어 넣지 않음 (다른 곳과 기준이 달라지는 것 방지).
"""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.marker import Marker

FONT_NAME = "맑은 고딕"
ACCENT = "44546A"
HEADER_FILL = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color=ACCENT)
SECTION_FONT = Font(name=FONT_NAME, size=11, bold=True, color=ACCENT)
LABEL_FONT = Font(name=FONT_NAME, size=10, bold=True)
NORMAL_FONT = Font(name=FONT_NAME, size=10)
NOTICE_FONT = Font(name=FONT_NAME, size=9, italic=True, color="808080")
THIN = Side(style="thin", color="D0D0D0")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CURRENCY_FMT = '#,##0"원";(#,##0)"원";"-"'
NUMBER_FMT = '#,##0;(#,##0);"-"'
PERCENT_FMT = '0.0%'


def _header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers, start=start_col):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER
    return row + 1


def _autosize(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _setup_print(ws):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _write_performance_sheet(wb: Workbook, data: dict):
    ws = wb.active
    ws.title = "판매성과리포트"

    ws["B2"] = data.get("period_card_label", "")
    ws["B2"].font = TITLE_FONT
    ws.merge_cells("B2:H2")

    # --- 대분류/중분류 Top5 ---
    r = 4
    ws.cell(row=r, column=2, value="대분류실적 Top5").font = SECTION_FONT
    ws.cell(row=r, column=6, value="중분류실적 Top5").font = SECTION_FONT
    r += 1
    header_row = r
    r = _header_row(ws, header_row, ["순위", "카테고리", "판매금액"], start_col=2)
    _header_row(ws, header_row, ["순위", "카테고리", "판매금액"], start_col=6)

    large5 = data.get("category_large_top5", [])
    mid5 = data.get("category_mid_top5", [])
    for i in range(5):
        row_idx = header_row + 1 + i
        if i < len(large5):
            item = large5[i]
            ws.cell(row=row_idx, column=2, value=item["rank"]).border = THIN_BORDER
            ws.cell(row=row_idx, column=3, value=item["name"]).border = THIN_BORDER
            c = ws.cell(row=row_idx, column=4, value=item["sales_amount"])
            c.number_format = CURRENCY_FMT
            c.border = THIN_BORDER
        if i < len(mid5):
            item = mid5[i]
            ws.cell(row=row_idx, column=6, value=item["rank"]).border = THIN_BORDER
            ws.cell(row=row_idx, column=7, value=item["name"]).border = THIN_BORDER
            c = ws.cell(row=row_idx, column=8, value=item["sales_amount"])
            c.number_format = CURRENCY_FMT
            c.border = THIN_BORDER
    r = header_row + 7

    # --- 베스트10 ---
    ws.cell(row=r, column=2, value="베스트 10").font = SECTION_FONT
    r += 1
    best_header_row = r
    r = _header_row(ws, best_header_row,
                     ["순위", "SKU", "대분류", "중분류", "품명", "가용재고", "판매율", "클레임"],
                     start_col=2)
    best10 = data.get("best10", [])
    for i, item in enumerate(best10):
        row_idx = best_header_row + 1 + i
        vals = [item["rank"], item["sku"], item["category_large"], item["category_mid"],
                item["item_name"], item["available_stock"], item["sell_through_rate"], item["claims"]]
        for j, v in enumerate(vals, start=2):
            c = ws.cell(row=row_idx, column=j, value=v)
            c.border = THIN_BORDER
            c.font = NORMAL_FONT
            if j == 8:  # 판매율(0-based 8번째 = 리스트 인덱스 6 = sell_through_rate)
                c.number_format = PERCENT_FMT
    r = best_header_row + len(best10) + 2

    # --- 해당기간 vs 비교기간 이중꺾은선 차트 (네이티브) ---
    chart_title_row = r
    ws.cell(row=chart_title_row, column=2, value="해당기간 vs 비교기간 매출 추이").font = SECTION_FONT
    r += 1
    trend_header_row = r
    _header_row(ws, trend_header_row, ["구간", "해당기간", "비교기간"], start_col=2)

    current_trend = data.get("current_trend", [])
    compare_trend = data.get("compare_trend", [])
    n_points = max(len(current_trend), len(compare_trend))

    data_start_row = trend_header_row + 1
    for i in range(n_points):
        row_idx = data_start_row + i
        label = current_trend[i]["label"] if i < len(current_trend) else (
            compare_trend[i]["label"] if i < len(compare_trend) else f"{i+1}"
        )
        ws.cell(row=row_idx, column=2, value=label).border = THIN_BORDER
        cur_val = current_trend[i]["value"] if i < len(current_trend) else None
        cmp_val = compare_trend[i]["value"] if i < len(compare_trend) else None
        c1 = ws.cell(row=row_idx, column=3, value=cur_val)
        c1.number_format = CURRENCY_FMT
        c1.border = THIN_BORDER
        c2 = ws.cell(row=row_idx, column=4, value=cmp_val)
        c2.number_format = CURRENCY_FMT
        c2.border = THIN_BORDER
    data_end_row = data_start_row + n_points - 1

    if n_points > 0:
        chart = LineChart()
        chart.title = "해당기간 vs 비교기간 매출 추이"
        chart.style = 2
        chart.y_axis.title = "매출액(원)"
        chart.x_axis.title = "구간"
        chart.height = 9
        chart.width = 20

        cats = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row)
        cur_ref = Reference(ws, min_col=3, min_row=trend_header_row, max_row=data_end_row)
        cmp_ref = Reference(ws, min_col=4, min_row=trend_header_row, max_row=data_end_row)
        chart.add_data(cur_ref, titles_from_data=True)
        chart.add_data(cmp_ref, titles_from_data=True)
        chart.set_categories(cats)

        for series in chart.series:
            series.marker = Marker(symbol="circle", size=5)
            series.smooth = False

        ws.add_chart(chart, f"F{trend_header_row}")

    _autosize(ws, {"B": 10, "C": 22, "D": 16, "E": 16, "F": 10, "G": 22, "H": 16})
    _setup_print(ws)


def _write_item_detail_sheet(wb: Workbook, sheet_name: str, items: list):
    ws = wb.create_sheet(sheet_name)
    columns = ["대분류", "중분류", "SKU번호", "판매순위", "품명", "정상가",
               "판매금액(합산)", "비중", "전체비중", "가용재고", "판매율", "클레임", "입고예정일"]
    header_row = 1
    r = _header_row(ws, header_row, columns, start_col=1)

    first_data_row = r
    for item in items:
        row = [
            item["category_large"], item["category_mid"], item["sku"], item["rank"], item["item_name"],
            item["regular_price"], item["sales_amount"], None, None,
            item["available_stock"], item["sell_through_rate"], item["claims"], item["expected_arrival_date"],
        ]
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = THIN_BORDER
            if j == 6 or j == 7:
                c.number_format = CURRENCY_FMT
            if j == 11:
                c.number_format = PERCENT_FMT
        r += 1
    last_data_row = r - 1

    if last_data_row >= first_data_row:
        for row_idx in range(first_data_row, last_data_row + 1):
            share_formula = (
                f"=IFERROR(G{row_idx}/SUMIF(A${first_data_row}:A${last_data_row},"
                f"A{row_idx},G${first_data_row}:G${last_data_row}),\"-\")"
            )
            c1 = ws.cell(row=row_idx, column=8, value=share_formula)
            c1.number_format = PERCENT_FMT
            c1.border = THIN_BORDER
            total_share_formula = f"=IFERROR(G{row_idx}/SUM(G${first_data_row}:G${last_data_row}),\"-\")"
            c2 = ws.cell(row=row_idx, column=9, value=total_share_formula)
            c2.number_format = PERCENT_FMT
            c2.border = THIN_BORDER

        total_row = last_data_row + 1
        ws.cell(row=total_row, column=5, value="합계").font = LABEL_FONT
        c = ws.cell(row=total_row, column=7, value=f"=SUM(G{first_data_row}:G{last_data_row})")
        c.number_format = CURRENCY_FMT
        c.font = LABEL_FONT

    ws.cell(row=last_data_row + 3, column=1,
            value="※ 정상/취소/반품/교환 등 모든 주문상태를 포함한 판매금액 합산입니다.").font = NOTICE_FONT

    _autosize(ws, {"A": 12, "B": 12, "C": 14, "D": 10, "E": 22, "F": 12,
                    "G": 16, "H": 10, "I": 10, "J": 12, "K": 10, "L": 10, "M": 14})
    _setup_print(ws)


PERIOD_TYPE_LABEL = {
    "daily": "일간", "weekly": "주간", "monthly": "월간",
    "quarterly": "분기", "yearly": "연간",
}


def generate_excel_report(data: dict, output_dir: str = "/tmp/reports") -> str:
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    _write_performance_sheet(wb, data)
    _write_item_detail_sheet(wb, "해당기간 판매 상세", data.get("current_period_items", []))
    _write_item_detail_sheet(wb, "비교기간 판매 상세", data.get("comparison_period_items", []))

    # 파일명 규칙: "YYYY.MM.DD_주간매출보고서.xlsx" (날짜는 기간의 시작일 기준)
    period_meta = data.get("period_meta", {})
    start_date_str = period_meta.get("start_date", "")  # "YYYY-MM-DD"
    period_type = period_meta.get("period_type", "")
    label = PERIOD_TYPE_LABEL.get(period_type, "")

    if start_date_str:
        date_part = start_date_str.replace("-", ".")
    else:
        date_part = datetime.now().strftime("%Y.%m.%d")

    filename = f"{date_part}_{label}매출보고서.xlsx" if label else f"{date_part}_매출보고서.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath